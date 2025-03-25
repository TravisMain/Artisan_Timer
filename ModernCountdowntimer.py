import tkinter as tk
from tkinter import messagebox
import time
import pickle
import os
import threading
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

class ModernCountdownTimer:
    def __init__(self, root):
        self.root = root
        self.root.title("Modern Countdown Timer")
        self.root.geometry("500x499")
        self.root.resizable(False, False)
        self.root.configure(bg="#2E3440")
        
        # Set the window icon
        try:
            icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "timer.ico")
            self.root.iconbitmap(icon_path)
        except Exception as e:
            error_log = os.path.join(os.path.expanduser("~"), "Downloads", "timer_error.log")
            with open(error_log, "a") as f:
                f.write(f"{datetime.now()}: Failed to load icon: {str(e)}\n")
        
        # Timer variables
        self.hours = 0
        self.minutes = 0
        self.seconds = 0
        self.total_seconds = 0
        self.remaining_seconds = 0
        self.running = False
        self.paused = False
        self.timer_thread = None
        self.last_pause_time = None
        
        # Mini window variables
        self.mini_window = None
        self.mini_time_display = None
        
        # File paths
        self.downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        self.save_file = os.path.join(self.downloads_dir, "timer_state.pkl")
        self.excel_file = None
        self.error_log = os.path.join(self.downloads_dir, "timer_error.log")
        
        # Job fields
        self.job_number = ""
        self.job_description = ""
        
        # Build number
        self.build_number = "1.0.0"
        
        # Create UI elements
        self.create_widgets()
        
        # Load saved state
        self.load_state()
        if os.path.exists(self.save_file):
            if messagebox.askyesno("Resume Timer", "A previous timer session was found. Would you like to resume?"):
                self.resume_timer()
            else:
                self.reset_timer()
        
        # Window protocols
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.bind("<Unmap>", self.on_minimize)
        self.root.bind("<Map>", self.on_restore)
        
    def create_widgets(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('TFrame', background='#2E3440')
        self.style.configure('TLabel', background='#2E3440', foreground='#ECEFF4', font=('Segoe UI', 10))
        self.style.configure('TButton', 
                             background='#5E81AC', 
                             foreground='#ECEFF4', 
                             borderwidth=0,
                             focusthickness=3,
                             focuscolor='#88C0D0',
                             font=('Segoe UI', 10, 'bold'))
        self.style.map('TButton', 
                       background=[('active', '#81A1C1'), ('disabled', '#4C566A')],
                       foreground=[('disabled', '#D8DEE9')])
        
        self.title_frame = ttk.Frame(self.root)
        self.title_frame.pack(pady=(20, 10))
        title_label = ttk.Label(self.title_frame, text="COUNTDOWN TIMER", font=('Segoe UI', 16, 'bold'))
        title_label.pack()
        
        self.input_frame = ttk.Frame(self.root)
        self.input_frame.pack(pady=15, fill=tk.X)
        
        time_frame = ttk.Frame(self.input_frame)
        time_frame.pack(pady=(0, 10))
        
        hours_frame = ttk.Frame(time_frame)
        hours_frame.grid(row=0, column=0, padx=10)
        self.hour_var = tk.StringVar(value="0")
        self.hour_spinbox = tk.Spinbox(hours_frame, from_=0, to=999, wrap=True, textvariable=self.hour_var, width=3, font=('Segoe UI', 24), bg="#3B4252", fg="#ECEFF4", buttonbackground="#4C566A", relief=tk.FLAT, justify=tk.CENTER)
        self.hour_spinbox.pack()
        ttk.Label(hours_frame, text="Hours", font=('Segoe UI', 8)).pack()
        
        separator1 = ttk.Label(time_frame, text=":", font=('Segoe UI', 24, 'bold'))
        separator1.grid(row=0, column=1)
        
        minutes_frame = ttk.Frame(time_frame)
        minutes_frame.grid(row=0, column=2, padx=10)
        self.minute_var = tk.StringVar(value="0")
        self.minute_spinbox = tk.Spinbox(minutes_frame, from_=0, to=59, wrap=True, textvariable=self.minute_var, width=3, font=('Segoe UI', 24), bg="#3B4252", fg="#ECEFF4", buttonbackground="#4C566A", relief=tk.FLAT, justify=tk.CENTER)
        self.minute_spinbox.pack()
        ttk.Label(minutes_frame, text="Minutes", font=('Segoe UI', 8)).pack()
        
        separator2 = ttk.Label(time_frame, text=":", font=('Segoe UI', 24, 'bold'))
        separator2.grid(row=0, column=3)
        
        seconds_frame = ttk.Frame(time_frame)
        seconds_frame.grid(row=0, column=4, padx=10)
        self.second_var = tk.StringVar(value="0")
        self.second_spinbox = tk.Spinbox(seconds_frame, from_=0, to=59, wrap=True, textvariable=self.second_var, width=3, font=('Segoe UI', 24), bg="#3B4252", fg="#ECEFF4", buttonbackground="#4C566A", relief=tk.FLAT, justify=tk.CENTER)
        self.second_spinbox.pack()
        ttk.Label(seconds_frame, text="Seconds", font=('Segoe UI', 8)).pack()
        
        self.job_frame = ttk.Frame(self.input_frame)
        self.job_frame.pack(pady=10)
        
        ttk.Label(self.job_frame, text="Job Number:", font=('Segoe UI', 10)).grid(row=0, column=0, padx=5, sticky="e")
        self.job_number_entry = ttk.Entry(self.job_frame, width=20, font=('Segoe UI', 10))
        self.job_number_entry.grid(row=0, column=1, padx=5)
        
        ttk.Label(self.job_frame, text="Job Description:", font=('Segoe UI', 10)).grid(row=1, column=0, padx=5, sticky="e")
        self.job_description_entry = ttk.Entry(self.job_frame, width=20, font=('Segoe UI', 10))
        self.job_description_entry.grid(row=1, column=1, padx=5)
        
        self.display_frame = ttk.Frame(self.root)
        self.display_frame.pack(pady=15)
        self.time_display_frame = tk.Frame(self.display_frame, bg="#434C5E", padx=20, pady=10, relief=tk.RAISED, bd=0)
        self.time_display_frame.pack()
        self.time_display = ttk.Label(self.time_display_frame, text="000:00:00", font=('Segoe UI', 36, 'bold'), foreground="#88C0D0", background="#434C5E")
        self.time_display.pack()
        
        self.button_frame = ttk.Frame(self.root)
        self.button_frame.pack(pady=20)
        
        self.start_button = ttk.Button(self.button_frame, text="Start", width=12, command=self.start_timer, style='TButton')
        self.start_button.grid(row=0, column=0, padx=8)
        
        self.pause_button = ttk.Button(self.button_frame, text="Pause", width=12, command=self.pause_timer, state=tk.DISABLED, style='TButton')
        self.pause_button.grid(row=0, column=1, padx=8)
        
        self.reset_button = ttk.Button(self.button_frame, text="Reset", width=12, command=self.confirm_reset, state=tk.DISABLED, style='TButton')
        self.reset_button.grid(row=0, column=2, padx=8)
        
        copyright_frame = ttk.Frame(self.root)
        copyright_frame.pack(side=tk.BOTTOM, pady=10)
        copyright_label = ttk.Label(copyright_frame, text=f"\u00A9 GuthSouth Africa - Build {self.build_number}", font=('Segoe UI', 8), foreground="#ECEFF4")
        copyright_label.pack()
        
    def start_timer(self):
        if not self.running:
            if not self.paused:
                try:
                    self.hours = int(self.hour_var.get())
                    self.minutes = int(self.minute_var.get())
                    self.seconds = int(self.second_var.get())
                    
                    if self.hours < 0 or self.minutes < 0 or self.seconds < 0:
                        messagebox.showerror("Invalid Input", "Please enter positive values.")
                        return
                    
                    if self.minutes >= 60 or self.seconds >= 60:
                        messagebox.showerror("Invalid Input", "Minutes and seconds must be less than 60.")
                        return
                    
                    self.total_seconds = self.hours * 3600 + self.minutes * 60 + self.seconds
                    if self.total_seconds == 0:
                        messagebox.showerror("Invalid Input", "Please enter a time greater than zero.")
                        return
                    
                    self.remaining_seconds = self.total_seconds
                    
                    self.job_number = self.job_number_entry.get().strip()
                    self.job_description = self.job_description_entry.get().strip()
                    if not self.job_number or not self.job_description:
                        messagebox.showerror("Invalid Input", "Please enter both Job Number and Job Description.")
                        return
                    self.excel_file = os.path.join(self.downloads_dir, f"{self.job_number}_{self.job_description}.xlsx")
                    with open(self.error_log, "a") as f:
                        f.write(f"{datetime.now()}: Excel file set to {self.excel_file}\n")
                    
                except ValueError:
                    messagebox.showerror("Invalid Input", "Please enter valid numbers for time.")
                    return
            
            self.running = True
            self.paused = False
            self.start_button.config(state=tk.DISABLED)
            self.pause_button.config(state=tk.NORMAL, text="Pause")
            self.reset_button.config(state=tk.NORMAL)
            
            self.input_frame.pack_forget()
            
            self.hour_spinbox.config(state=tk.DISABLED)
            self.minute_spinbox.config(state=tk.DISABLED)
            self.second_spinbox.config(state=tk.DISABLED)
            self.job_number_entry.config(state=tk.DISABLED)
            self.job_description_entry.config(state=tk.DISABLED)
            
            self.timer_thread = threading.Thread(target=self.update_timer)
            self.timer_thread.daemon = True
            self.timer_thread.start()
    
    def update_timer(self):
        while self.running and self.remaining_seconds > 0:
            if not self.paused:
                mins, secs = divmod(self.remaining_seconds, 60)
                hours, mins = divmod(mins, 60)
                time_str = f"{hours:03d}:{mins:02d}:{secs:02d}"
                self.root.after(0, lambda: self.time_display.config(text=time_str))
                
                if self.mini_window and self.mini_time_display:
                    self.root.after(0, lambda: self.mini_time_display.config(text=time_str))
                
                if self.remaining_seconds <= 10:
                    self.root.after(0, lambda: self.time_display.config(foreground="#BF616A"))
                    if self.mini_time_display:
                        self.root.after(0, lambda: self.mini_time_display.config(foreground="#BF616A"))
                elif self.remaining_seconds <= 60:
                    self.root.after(0, lambda: self.time_display.config(foreground="#EBCB8B"))
                    if self.mini_time_display:
                        self.root.after(0, lambda: self.mini_time_display.config(foreground="#EBCB8B"))
                
                time.sleep(1)
                self.remaining_seconds -= 1
                
                if self.remaining_seconds % 10 == 0:
                    self.save_state()
            else:
                time.sleep(0.1)
        
        if self.running and self.remaining_seconds <= 0:
            self.root.after(0, self.timer_complete)
    
    def timer_complete(self):
        self.time_display.config(text="000:00:00", foreground="#A3BE8C")
        if self.mini_time_display:
            self.mini_time_display.config(text="000:00:00", foreground="#A3BE8C")
        messagebox.showinfo("Timer Complete", "The countdown timer has finished!")
        self.reset_timer()
    
    def pause_timer(self):
        if self.running:
            if not self.paused:
                self.paused = True
                self.pause_button.config(text="Resume")
                self.last_pause_time = datetime.now()
                self.save_state()
                self.log_to_excel("Paused")
            else:
                self.paused = False
                self.pause_button.config(text="Pause")
                self.log_to_excel("Resumed")
    
    def log_to_excel(self, state):
        mins, secs = divmod(self.remaining_seconds, 60)
        hours, mins = divmod(mins, 60)
        remaining_time = f"{hours:03d}:{mins:02d}:{secs:02d}"
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        pause_duration = ""
        if state == "Resumed" and self.last_pause_time:
            duration = datetime.now() - self.last_pause_time
            seconds = int(duration.total_seconds())
            mins, secs = divmod(seconds, 60)
            hours, mins = divmod(mins, 60)
            pause_duration = f"{hours:02d}:{mins:02d}:{secs:02d}"
            self.last_pause_time = None
        
        data = [self.job_number, self.job_description, state, current_time, remaining_time, pause_duration]
        
        try:
            with open(self.error_log, "a") as f:
                f.write(f"{datetime.now()}: Attempting to save to {self.excel_file}\n")
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["Job Number", "Job Description", "State", "DateTime", "Remaining Time", "Pause Duration"])
            
            ws.append(data)
            wb.save(self.excel_file)
            with open(self.error_log, "a") as f:
                f.write(f"{datetime.now()}: Successfully saved to {self.excel_file}\n")
        except Exception as e:
            with open(self.error_log, "a") as f:
                f.write(f"{datetime.now()}: Failed to save Excel: {str(e)}\n")
            messagebox.showerror("Excel Error", f"Failed to save to Excel: {str(e)}")
    
    def confirm_reset(self):
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset the timer?"):
            self.reset_timer()
    
    def reset_timer(self):
        self.running = False
        self.paused = False
        self.remaining_seconds = 0
        self.last_pause_time = None
        
        if self.timer_thread and self.timer_thread.is_alive():
            self.timer_thread.join(timeout=0.1)
        
        self.time_display.config(text="000:00:00", foreground="#88C0D0")
        self.start_button.config(state=tk.NORMAL)
        self.pause_button.config(state=tk.DISABLED, text="Pause")
        self.reset_button.config(state=tk.DISABLED)
        
        self.input_frame.pack(pady=15, fill=tk.X, before=self.display_frame)
        
        self.hour_spinbox.config(state=tk.NORMAL)
        self.minute_spinbox.config(state=tk.NORMAL)
        self.second_spinbox.config(state=tk.NORMAL)
        self.job_number_entry.config(state=tk.NORMAL)
        self.job_description_entry.config(state=tk.NORMAL)
        
        self.hour_var.set("0")
        self.minute_var.set("0")
        self.second_var.set("0")
        self.job_number_entry.delete(0, tk.END)
        self.job_description_entry.delete(0, tk.END)
        
        self.close_mini_window()
        self.clear_saved_state()
        self.excel_file = None
    
    def resume_timer(self):
        if self.remaining_seconds > 0:
            self.running = True
            self.paused = False
            self.start_button.config(state=tk.DISABLED)
            self.pause_button.config(state=tk.NORMAL, text="Pause")
            self.reset_button.config(state=tk.NORMAL)
            
            self.input_frame.pack_forget()
            
            self.hour_spinbox.config(state=tk.DISABLED)
            self.minute_spinbox.config(state=tk.DISABLED)
            self.second_spinbox.config(state=tk.DISABLED)
            self.job_number_entry.config(state=tk.DISABLED)
            self.job_description_entry.config(state=tk.DISABLED)
            
            self.timer_thread = threading.Thread(target=self.update_timer)
            self.timer_thread.daemon = True
            self.timer_thread.start()

    def save_state(self):
        state = {
            'remaining_seconds': self.remaining_seconds,
            'running': self.running,
            'paused': self.paused,
            'hours': self.hours,
            'minutes': self.minutes,
            'seconds': self.seconds,
            'total_seconds': self.total_seconds,
            'job_number': self.job_number,
            'job_description': self.job_description,
            'excel_file': self.excel_file
        }
        try:
            with open(self.save_file, 'wb') as f:
                pickle.dump(state, f)
        except Exception as e:
            with open(self.error_log, "a") as f:
                f.write(f"{datetime.now()}: Failed to save state: {str(e)}\n")

    def load_state(self):
        if os.path.exists(self.save_file):
            try:
                with open(self.save_file, 'rb') as f:
                    state = pickle.load(f)
                    self.remaining_seconds = state['remaining_seconds']
                    self.running = False
                    self.paused = state['paused']
                    self.hours = state['hours']
                    self.minutes = state['minutes']
                    self.seconds = state['seconds']
                    self.total_seconds = state['total_seconds']
                    self.job_number = state.get('job_number', '')
                    self.job_description = state.get('job_description', '')
                    self.excel_file = state.get('excel_file', None)
                    
                    mins, secs = divmod(self.remaining_seconds, 60)
                    hours, mins = divmod(mins, 60)
                    self.time_display.config(text=f"{hours:03d}:{mins:02d}:{secs:02d}")
                    
                    self.hour_var.set(str(self.hours))
                    self.minute_var.set(str(self.minutes))
                    self.second_var.set(str(self.seconds))
                    self.job_number_entry.delete(0, tk.END)
                    self.job_number_entry.insert(0, self.job_number)
                    self.job_description_entry.delete(0, tk.END)
                    self.job_description_entry.insert(0, self.job_description)
            except:
                self.clear_saved_state()
        else:
            self.time_display.config(text="000:00:00")
            self.hour_var.set("0")
            self.minute_var.set("0")
            self.second_var.set("0")
            self.job_number_entry.delete(0, tk.END)
            self.job_description_entry.delete(0, tk.END)

    def clear_saved_state(self):
        if os.path.exists(self.save_file):
            try:
                os.remove(self.save_file)
            except:
                pass

    def on_close(self):
        self.running = False
        if self.timer_thread and self.timer_thread.is_alive():
            self.timer_thread.join(timeout=0.1)
        self.close_mini_window()
        self.save_state()
        self.root.destroy()

    def on_minimize(self, event):
        if self.running and self.root.state() == 'iconic':
            self.create_mini_window()

    def on_restore(self, event):
        if self.mini_window:
            self.close_mini_window()

    def create_mini_window(self):
        if not self.mini_window:
            self.mini_window = tk.Toplevel(self.root)
            self.mini_window.overrideredirect(True)
            self.mini_window.geometry("150x50+{}+0".format(self.root.winfo_screenwidth() - 150))
            self.mini_window.configure(bg="#434C5E")
            
            self.mini_time_display = ttk.Label(self.mini_window, text=self.time_display.cget("text"), font=('Segoe UI', 20, 'bold'), foreground="#88C0D0", background="#434C5E")
            self.mini_time_display.pack(expand=True)
            
            self.mini_window.attributes('-topmost', True)

    def close_mini_window(self):
        if self.mini_window:
            self.mini_window.destroy()
            self.mini_window = None
            self.mini_time_display = None

if __name__ == "__main__":
    root = tk.Tk()
    app = ModernCountdownTimer(root)
    root.mainloop()